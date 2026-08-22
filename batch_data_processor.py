#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
SCRIPT 1: BATCH DATA PROCESSOR (LỌC & TÍNH TOÁN DỮ LIỆU CẤN TRỪ TỒN KHO)
================================================================================
Phiên bản: Pure Python + openpyxl (KHÔNG CẦN PANDAS, KHÔNG CẦN EXCEL, CHẠY CỰC NHẸ)
Nhiệm vụ:
  1. Đọc file nguồn 'Stock Balance With Batch.xlsx'.
  2. Lọc các dòng âm tại Kho 50 và Kho 62.
  3. Lọc nguồn tồn dương tại Kho 61 và Kho 01 (theo đúng Stock Code).
  4. Lựa chọn các batch có số lượng nhỏ nhất trước (Smallest Qty First).
  5. Cộng dồn các batch nguyên vẹn cho đến khi đủ số lượng âm cần trừ.
     Nếu batch tiếp theo phải tách lẻ thì tạm bỏ qua (không tách batch).
  6. Xuất ra:
     - auto_input_queue.json (Dữ liệu hàng đợi cho Auto-Typer GUI)
     - auto_input_queue.xlsx (Bảng Excel chi tiết cấn trừ)
     - report_unmatched_shortage.xlsx (Báo cáo danh sách chưa cấn trừ xong)
================================================================================
"""

import sys
import os
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def format_excel_sheet(ws, header_color="1F4E78"):
    """Định dạng bảng tính Excel đẹp mắt với màu tiêu đề và căn lề"""
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=9)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def process_stock_data(input_file="Stock Balance With Batch.xlsx"):
    log(f"Bắt đầu xử lý file dữ liệu: {input_file}")
    if not os.path.exists(input_file):
        log(f"LỖI: Không tìm thấy file '{input_file}'. Vui lòng kiểm tra lại đường dẫn.")
        return False

    # 1. Đọc dữ liệu từ file Excel
    log("Đang đọc và phân tích cấu trúc bảng tính...")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        log(f"Lỗi khi mở file Excel: {e}")
        return False

    neg_dict = {}  # (stock_code, target_wh) -> total_negative_qty
    pos_dict = {}  # stock_code -> list of {'batch', 'wh', 'qty', 'bin'}
    neg_count = 0
    pos_count = 0
    neg_total_sum = 0
    pos_total_sum = 0

    # Duyệt từng dòng dữ liệu (bỏ 5 dòng đầu là header tiêu đề báo cáo)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        
        sc = str(row[0] or '').strip()
        wh = str(row[1] or '').strip()
        if wh in ['1', '1.0']:
            wh = '01'
        batch = str(row[4] or '').strip()
        bin_val = str(row[5] or '').strip()
        raw_qty = row[6]

        if not sc or raw_qty is None:
            continue

        try:
            qty = float(raw_qty)
        except:
            continue

        # 2. Phân loại nhóm Âm và Dương theo quy tắc nghiệp vụ
        # Nhóm Âm: Kho 50 và Kho 62, Qty < 0
        if qty < 0 and wh in ['50', '62']:
            neg_count += 1
            neg_total_sum += abs(qty)
            key = (sc, wh)
            neg_dict[key] = neg_dict.get(key, 0.0) + abs(qty)

        # Nhóm Dương: Kho 61 và Kho 01, Qty > 0
        elif qty > 0 and wh in ['61', '01']:
            pos_count += 1
            pos_total_sum += qty
            if sc not in pos_dict:
                pos_dict[sc] = []
            pos_dict[sc].append({
                'batch': batch,
                'wh': wh,
                'qty': int(qty),
                'bin': bin_val or '01'
            })

    log(f"-> Tổng số dòng âm (Kho 50 & 62): {neg_count:,} dòng | Tổng số lượng âm: {neg_total_sum:,.0f}")
    log(f"-> Tổng số dòng dương khả dụng (Kho 61 & 01): {pos_count:,} dòng | Tổng tồn dương: {pos_total_sum:,.0f}")

    # 3. Gom nhóm và thực hiện thuật toán cấn trừ (Smallest Batch First, No Partial Splitting)
    queue_actions = []
    skipped_records = []
    step_counter = 1

    for (sc, target_wh) in sorted(neg_dict.keys()):
        target_qty = int(neg_dict[(sc, target_wh)])
        # Sắp xếp các batch dương theo số lượng nhỏ nhất trước
        cand_pos = sorted(pos_dict.get(sc, []), key=lambda x: x['qty'])

        cur_sum = 0
        used_indices = []

        for idx, p_item in enumerate(cand_pos):
            b_qty = p_item['qty']
            if cur_sum + b_qty <= target_qty:
                cur_sum += b_qty
                used_indices.append(idx)
                queue_actions.append({
                    "step": step_counter,
                    "stock_code": sc,
                    "target_warehouse": target_wh,
                    "batch": p_item['batch'],
                    "source_warehouse": p_item['wh'],
                    "qty": b_qty,
                    "bin": "01",
                    "status": "READY"
                })
                step_counter += 1
            else:
                # Nếu lấy batch này sẽ bị vượt số âm (cần tách lẻ) -> Theo quy tắc: TẠM BỎ QUA
                pass

        # Cập nhật danh sách batch còn lại của stock code này
        pos_dict[sc] = [item for i_pos, item in enumerate(cand_pos) if i_pos not in used_indices]

        rem_missing = target_qty - cur_sum
        if rem_missing > 0:
            skipped_records.append({
                "stock_code": sc,
                "target_wh": target_wh,
                "total_neg": target_qty,
                "offset_done": cur_sum,
                "remaining": rem_missing,
                "ratio": round(cur_sum / target_qty * 100, 1) if target_qty > 0 else 0,
                "reason": "Các lô dương còn lại có số lượng lớn hơn phần thiếu (Cần tách lẻ) hoặc hết tồn"
            })

    total_offset_qty = sum(a['qty'] for a in queue_actions)
    total_neg_qty = int(neg_total_sum)

    log(f"-> Đã tạo thành công: {len(queue_actions)} lượt nhập liệu hợp lệ")
    log(f"-> Tổng số lượng giải quyết ngay: {total_offset_qty:,.0f} / {total_neg_qty:,.0f} ({total_offset_qty/total_neg_qty*100:.1f}%)")
    log(f"-> Số mã vật tư còn phần thiếu cần tách lẻ/theo dõi: {len(skipped_records)}")

    # 4. Xuất file auto_input_queue.json
    output_json = {
        "metadata": {
            "created_at": datetime.now().isoformat(),
            "source_file": input_file,
            "total_actions": int(len(queue_actions)),
            "total_offset_qty": int(total_offset_qty),
            "total_negative_qty": int(total_neg_qty),
            "offset_percentage": float(round(total_offset_qty / total_neg_qty * 100, 2)) if total_neg_qty > 0 else 0.0
        },
        "actions": queue_actions
    }
    with open("auto_input_queue.json", "w", encoding="utf-8") as f:
        json.dump(output_json, f, ensure_ascii=False, indent=2)
    log(" Đã xuất file hàng đợi: auto_input_queue.json")

    # 5. Xuất file auto_input_queue.xlsx bằng pure openpyxl
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Danh Sach Nhap Lieu"
    
    headers = [
        "STT", "Mã Vật Tư (Stock Code)", "Mã Lô Dương (Batch)", 
        "Kho Nguồn", "Kho Cấn Trừ (Target WH)", "Số Lượng Cấn Trừ (Qty)", 
        "Vị Trí Kệ (BIN)", "Quy Trình Gõ Form iScala"
    ]
    ws_out.append(headers)
    
    for item in queue_actions:
        ws_out.append([
            item["step"],
            item["stock_code"],
            item["batch"],
            item["source_warehouse"],
            item["target_warehouse"],
            item["qty"],
            item["bin"],
            f"Gõ [{item['batch']}] -> Enter -> Gõ [{item['target_warehouse']}] -> Enter -> Gõ [{item['qty']}] -> Enter -> Gõ [01] -> Enter x 4"
        ])
    
    format_excel_sheet(ws_out, header_color="1F4E78")
    wb_out.save("auto_input_queue.xlsx")
    log(" Đã xuất file bảng tính: auto_input_queue.xlsx")

    # 6. Xuất file report_unmatched_shortage.xlsx bằng pure openpyxl
    wb_rep = openpyxl.Workbook()
    
    # Sheet 1: Tổng Quan
    ws_sum = wb_rep.active
    ws_sum.title = "Tong Quan"
    ws_sum.append(["Chỉ Tiêu Thống Kê", "Giá Trị"])
    ws_sum.append(["Tổng số dòng âm tại Kho 50 & 62", f"{neg_count:,} dòng"])
    ws_sum.append(["Tổng số lượng âm cần xử lý", f"{total_neg_qty:,.0f}"])
    ws_sum.append(["Số lượt nhập tự động (Lô nguyên vẹn)", f"{len(queue_actions)} lượt"])
    ws_sum.append(["Số lượng âm đã giải quyết ngay", f"{total_offset_qty:,.0f}"])
    ws_sum.append(["Số lượng còn thiếu (Chờ tách lẻ / nhập thêm)", f"{total_neg_qty - total_offset_qty:,.0f}"])
    ws_sum.append(["Tỷ lệ cấn trừ ngay không cần tách lẻ", f"{total_offset_qty/total_neg_qty*100:.2f}%"])
    format_excel_sheet(ws_sum, header_color="2E75B6")

    # Sheet 2: Danh sách chờ tách lẻ
    ws_skip = wb_rep.create_sheet(title="Danh Sach Cho Tach Le")
    ws_skip.append([
        "Mã Vật Tư (Stock Code)", "Kho Bị Âm", "Tổng Số Lượng Âm", 
        "Đã Bù Trọn Vẹn", "Còn Thiếu (Chờ Tách Lẻ)", "Tỷ Lệ Bù (%)", "Lý Do"
    ])
    for s in skipped_records:
        ws_skip.append([
            s["stock_code"],
            s["target_wh"],
            s["total_neg"],
            s["offset_done"],
            s["remaining"],
            s["ratio"],
            s["reason"]
        ])
    format_excel_sheet(ws_skip, header_color="C00000")
    
    wb_rep.save("report_unmatched_shortage.xlsx")
    log(" Đã xuất file báo cáo: report_unmatched_shortage.xlsx")
    log("=== HOÀN TẤT XỬ LÝ DỮ LIỆU THÀNH CÔNG ===")
    return True

if __name__ == "__main__":
    file_target = sys.argv[1] if len(sys.argv) > 1 else "Stock Balance With Batch.xlsx"
    process_stock_data(file_target)
